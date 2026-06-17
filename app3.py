# ============================================================
# CONCILIADOR DE COTIZACIONES PDF  |  app.py  v5.0
# ============================================================
# Novedades v5.0 (sobre v4.3):
#   V1  — Cota de Cordura Proporcional: si el "total" hallado
#         es >1.5× el subtotal o la suma de líneas, se descarta.
#   V2  — Blacklist ampliada (_BUDGET_EXCL): ignora líneas con
#         "Anexo N", "Aportante", "costo total", "inversión".
#   V3  — Fallback con validación proporcional: en Estrategia 4
#         también se filtra por la cota antes de usarlo.
#   V4  — import math; _safe_f usa math.isnan / math.isinf.
#   V5  — pdf_upload_hash: hash del conjunto de archivos subidos
#         (separa la lógica de "qué se subió" de "qué resultó"
#         al combinar, evitando una recomputación innecesaria).
#
# Características heredadas de v4.3:
#   M1  — Carga múltiple de PDFs (accept_multiple_files=True).
#   M2  — Selector de PDF activo por sección (páginas locales).
#   M3  — Indicador visual de archivo fuente por página.
#   M4  — Descarga individual y consolidada.
#   N1  — Botones de navegación sincronizados (nav_page_input).
#   N2  — Dark Mode con CSS variables.
#   N3  — Tipo de sección (Presupuesto Global / Cotización).
#   N4  — recalc_derived robusto e idempotente.
# ============================================================

from __future__ import annotations

import datetime
import hashlib
import io
import math
import re
from typing import Optional

import fitz          # PyMuPDF
import numpy as np
import openpyxl
import pandas as pd
import pdfplumber
import requests
import streamlit as st
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ─────────────────────────────────────────────────────────────
# CONFIGURACIÓN DE PÁGINA E INTERFAZ (CSS)
# ─────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Conciliador de Cotizaciones",
    page_icon="📋",
    layout="wide",
    initial_sidebar_state="expanded",
)

st.markdown(
    """
<style>
:root {
    color-scheme: light dark;
    --app-bg: #ffffff;
    --app-text: #1f1f24;
    --app-muted: #7a4a5a;
    --panel-bg: #fdf5f7;
    --panel-border: #e8b4c0;
    --accent: #6E152E;
    --accent-2: #a02048;
    --accent-soft: #f8ccd6;
    --sidebar-bg: #6E152E;
    --sidebar-text: #f0e0e6;
    --input-bg: #f5e0e6;
    --input-text: #1a1a1a;
    --file-tag-bg: #2c3e50;
    --file-tag-text: #ecf0f1;
}
@media (prefers-color-scheme: dark) {
    :root {
        --app-bg: #111217;
        --app-text: #f2edf0;
        --app-muted: #d8a8b7;
        --panel-bg: #1d1519;
        --panel-border: #5b2637;
        --accent: #b84a6b;
        --accent-2: #7d203e;
        --accent-soft: #ffd9e3;
        --sidebar-bg: #2a0f1a;
        --sidebar-text: #ffeaf0;
        --input-bg: #2b2026;
        --input-text: #f8f1f3;
        --file-tag-bg: #324355;
        --file-tag-text: #f1f5f9;
    }
}
[data-testid="stAppViewContainer"] { background: var(--app-bg); color: var(--app-text); }
[data-testid="stSidebar"] { background: var(--sidebar-bg); }
[data-testid="stSidebar"] * { color: var(--sidebar-text) !important; }
[data-testid="stSidebar"] input,
[data-testid="stSidebar"] select,
[data-testid="stSidebar"] textarea {
    color: var(--input-text) !important;
    background-color: var(--input-bg) !important;
    border-radius: 6px !important;
}
[data-testid="stSidebar"] input::placeholder,
[data-testid="stSidebar"] textarea::placeholder {
    color: var(--app-muted) !important; opacity: 1 !important;
}
.hdr {
    background: linear-gradient(90deg, var(--accent), var(--accent-2));
    padding: 14px 22px; border-radius: 8px; margin-bottom: 18px;
}
.hdr h1 { color: #fff; font-size: 1.4rem; margin: 0; font-weight: 700; }
.hdr p  { color: var(--accent-soft); font-size: .87rem; margin: 4px 0 0; }
.ptitle {
    background: var(--accent); color: #fff !important; font-weight: 700;
    font-size: .88rem; padding: 7px 14px; border-radius: 6px 6px 0 0; margin-bottom: 4px;
}
.kpi {
    background: var(--panel-bg); border: 1px solid var(--panel-border);
    border-radius: 8px; padding: 10px 12px; text-align: center; margin-bottom: 6px;
}
.kpi .v { font-size: 1.2rem; font-weight: 700; color: var(--accent); }
.kpi .l { font-size: .72rem; color: var(--app-muted); }
.tag-moneda {
    background: var(--accent); color: #fff !important; padding: 2px 8px;
    border-radius: 4px; font-size: .78rem; font-weight: 600;
}
.tag-file {
    background: var(--file-tag-bg); color: var(--file-tag-text) !important;
    padding: 2px 8px; border-radius: 4px; font-size: .75rem;
    font-weight: 600; margin-right: 4px;
}
</style>
""",
    unsafe_allow_html=True,
)


# ─────────────────────────────────────────────────────────────
# CONSTANTES Y REGEX
# ─────────────────────────────────────────────────────────────
NATIVE_MIN_CHARS_PER_PAGE = 80
MAX_PLAUSIBLE_MXN = 50_000_000

_BANXICO_SERIES = {"USD": "SF43718", "EUR": "SF46410", "CAD": "SF60653"}
_BANXICO_URL = "https://www.banxico.org.mx/SieAPIRest/service/v1"

# FIX-V5: columna "Tipo" incluida desde v4.3
_COLS = [
    "Tipo",
    "Fecha", "Rubro", "QT", "T. Cambio", "(+ IVA)", "Cantidad",
    "Precio Unitario", "Subtotal (Sin IVA)", "IVA 16%", "Total con IVA",
    "Diferencia final", "Monto en Anexo Escrito", "Observaciones",
]
_WIDTHS_COL = [20, 12, 52, 5, 10, 7, 8, 16, 18, 17, 16, 18, 22, 48]

_MONEY_RE = re.compile(
    r"\$\s*([\d,]+(?:\.\d{1,2})?)"
    r"|(?<!\d)([\d]{1,3}(?:,\d{3})+(?:\.\d{1,2})?)"
    r"|(?<!\d)(\d{1,9}\.\d{2})(?!\d)"
)
_MONETARY_CTX = re.compile(
    r"total|importe|monto|precio|valor|costo|cobro|cargo|pago"
    r"|subtotal|honorarios|tarifa|renta|fianza",
    re.IGNORECASE,
)
_DATE_RE = re.compile(
    r"(\d{1,2})\s*(?:de)?\s*([a-záéíóúñA-ZÁÉÍÓÚÑ]+)[,\s]*(?:del?\s*)?(\d{4})"
    r"|(\d{4})[\s.\-/]+(\d{1,2})[\s.\-/]+(\d{1,2})"
    r"|(\d{1,2})[\s.\-/]+(\d{1,2})[\s.\-/]+(\d{2,4})",
    re.IGNORECASE,
)
_ITEM_RE = re.compile(
    r"^\d+\s+(\d+)\s+\w+\s+(.+?)\s+([\d,]+\.\d{2})\s+([\d,]+\.\d{2})\s*$"
)
_ES_DE_RE = re.compile(r"es\s+de\s+\$?\s*([\d,]+(?:\.\d{1,2})?)", re.IGNORECASE)

# V2: Blacklist ampliada — evita montos del presupuesto global o de resúmenes de anexos
_BUDGET_EXCL = re.compile(
    r"presupuestal|presupuesto\s+total|costo\s+total|proyecto|inversi[óo]n"
    r"|anexo\s*\d|aportante",
    re.IGNORECASE,
)

_MESES = {
    "enero": 1, "febrero": 2, "marzo": 3, "abril": 4, "mayo": 5, "junio": 6,
    "julio": 7, "agosto": 8, "septiembre": 9, "octubre": 10, "noviembre": 11, "diciembre": 12,
    "ene": 1, "feb": 2, "mar": 3, "abr": 4, "may": 5, "jun": 6,
    "jul": 7, "ago": 8, "sep": 9, "oct": 10, "nov": 11, "dic": 12,
}

_FMT_MONEY = (
    '_-"$ "* #,##0.00_-;'
    '\\-"$ "* #,##0.00_-;'
    '_-"$ "* "-"??_-;_-@_-'
)
_FMT_DATE = "mm-dd-yy"


# ─────────────────────────────────────────────────────────────
# SESSION STATE
# ─────────────────────────────────────────────────────────────
_SS_DEFAULTS: dict = {
    "pdf_files":         [],
    "pdf_combined":      None,
    "pdf_combined_hash": None,
    # V5: hash del conjunto subido (distinto al del combinado)
    "pdf_upload_hash":   None,
    "total_pages":       0,
    "page_map":          [],
    "current_page":      0,
    "nav_page_input":    1,   # N1: sincronización de botones
    "num_sec":           1,
    "sec_cfg":           [],
    "df":                None,
    "editor_df":         None,
    "editor_version":    0,
    "df_hash":           "",
    "extracted":         False,
    "bx_cache":          {},
    "proyecto_nombre":   "",
    "_rerun_guard":      False,
}
for _k, _v in _SS_DEFAULTS.items():
    st.session_state.setdefault(_k, _v)


# ─────────────────────────────────────────────────────────────
# HELPERS
# ─────────────────────────────────────────────────────────────
def _md5(b: bytes) -> str:
    return hashlib.md5(b).hexdigest()


def _is_blank_value(v) -> bool:
    if v is None:
        return True
    try:
        if bool(pd.isna(v)):
            return True
    except (TypeError, ValueError):
        pass
    return isinstance(v, str) and not v.strip()


def _normalize_editor_df(df: pd.DataFrame | None) -> pd.DataFrame:
    if df is None:
        return pd.DataFrame(columns=_COLS)
    clean = pd.DataFrame(df).copy()
    private_cols = [c for c in clean.columns if str(c).startswith("_")]
    clean = clean.drop(columns=private_cols, errors="ignore").reindex(columns=_COLS)
    if clean.empty:
        return clean.reset_index(drop=True)
    editable_cols = [c for c in _COLS if c != "Diferencia final"]
    blank_rows = clean[editable_cols].apply(
        lambda row: all(_is_blank_value(v) for v in row), axis=1
    )
    return clean.loc[~blank_rows].reset_index(drop=True)


def _df_hash(df: pd.DataFrame) -> str:
    if df is None:
        return ""
    df = _normalize_editor_df(df)
    hash_df = df.astype(object).where(pd.notna(df), "")
    try:
        return hashlib.md5(
            pd.util.hash_pandas_object(hash_df, index=False).values.tobytes()
        ).hexdigest()
    except Exception:
        return hashlib.md5(hash_df.to_csv(index=False).encode()).hexdigest()


def _safe_f(v) -> Optional[float]:
    if v is None:
        return None
    try:
        # V4: usa math.isnan / math.isinf (más robusto que f != f)
        f = float(str(v).replace(",", "").strip())
        return None if math.isnan(f) or math.isinf(f) else f
    except (ValueError, TypeError):
        return None


def _money(txt: str) -> Optional[float]:
    for m in _MONEY_RE.finditer(str(txt)):
        raw = m.group(1) or m.group(2) or m.group(3)
        if raw:
            try:
                v = float(raw.replace(",", ""))
                if v > 0:
                    return v
            except ValueError:
                continue
    return None


def _date(text: str) -> Optional[datetime.date]:
    for m in _DATE_RE.finditer(text.lower()):
        g = m.groups()
        try:
            if g[0]:
                mes_str = g[1].lower().strip()
                mes_num = _MESES.get(mes_str) or _MESES.get(mes_str[:3])
                if not mes_num:
                    for k, v in _MESES.items():
                        if mes_str.startswith(k[:3]):
                            mes_num = v
                            break
                if mes_num:
                    return datetime.date(int(g[2]), mes_num, int(g[0]))
            if g[3]:
                return datetime.date(int(g[3]), int(g[4]), int(g[5]))
            if g[6]:
                yr = int(g[8])
                return datetime.date(yr + 2000 if yr < 100 else yr, int(g[7]), int(g[6]))
        except (ValueError, KeyError):
            continue
    return None


# ─────────────────────────────────────────────────────────────
# CONSOLIDACIÓN DE PDFs
# ─────────────────────────────────────────────────────────────
def _build_combined_pdf(file_list: list[dict]) -> tuple[bytes, str, int, list[dict]]:
    combined = fitz.open()
    page_map: list[dict] = []

    for file_idx, f_info in enumerate(file_list):
        src = fitz.open(stream=f_info["bytes"], filetype="pdf")
        n = len(src)
        combined.insert_pdf(src)
        for local_p in range(n):
            page_map.append({
                "file_idx": file_idx,
                "local_page": local_p,
                "name": f_info["name"],
            })
        src.close()

    buf = io.BytesIO()
    combined.save(buf)
    combined.close()
    combined_bytes = buf.getvalue()
    return combined_bytes, _md5(combined_bytes), len(page_map), page_map


# ─────────────────────────────────────────────────────────────
# OCR Y RENDER
# ─────────────────────────────────────────────────────────────
@st.cache_resource(show_spinner=False, max_entries=1)
def _get_ocr():
    try:
        from rapidocr_onnxruntime import RapidOCR
        return RapidOCR(det_model_dir=None, rec_model_dir=None, cls_model_dir=None)
    except Exception as exc:
        st.warning(f"RapidOCR no disponible ({exc}). OCR desactivado.")
        return None


def _ocr_page(pdf_bytes: bytes, idx: int) -> str:
    ocr = _get_ocr()
    if ocr is None:
        return ""
    try:
        with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
            pix = doc[idx].get_pixmap(
                matrix=fitz.Matrix(1.0, 1.0), colorspace=fitz.csRGB, alpha=False
            )
            img = np.frombuffer(pix.samples, np.uint8).reshape(
                pix.height, pix.width, 3
            )[:, :, ::-1]
            result, _ = ocr(img)
            if result:
                return "\n".join(r[1] for r in result if r and len(r) > 1)
            return ""
    except Exception as exc:
        st.warning(f"Error OCR pág. {idx + 1}: {exc}")
        return ""


@st.cache_data(max_entries=200, show_spinner=False)
def _render(pdf_hash: str, idx: int) -> bytes:
    """F3: pdf_bytes desde session_state; hash invalida la caché al cambiar."""
    pdf_bytes = st.session_state.pdf_combined
    with fitz.open(stream=pdf_bytes, filetype="pdf") as doc:
        pix = doc[idx].get_pixmap(
            matrix=fitz.Matrix(1.5, 1.5), colorspace=fitz.csRGB, alpha=False
        )
        return pix.tobytes("png")


# ─────────────────────────────────────────────────────────────
# API BANXICO
# ─────────────────────────────────────────────────────────────
def _banxico_tc(moneda: str, fecha: datetime.date, token: str) -> Optional[float]:
    if moneda not in _BANXICO_SERIES or not token:
        return None
    cache_key = (moneda, fecha.isoformat())
    if cache_key in st.session_state.bx_cache:
        return st.session_state.bx_cache[cache_key]

    serie = _BANXICO_SERIES[moneda]
    f_ini = (fecha - datetime.timedelta(days=5)).isoformat()
    url = f"{_BANXICO_URL}/series/{serie}/datos/{f_ini}/{fecha.isoformat()}"
    try:
        resp = requests.get(url, headers={"Bmx-Token": token, "Accept": "application/json"}, timeout=10)
        resp.raise_for_status()
        datos = resp.json()["bmx"]["series"][0]["datos"]
        if not datos:
            st.session_state.bx_cache[cache_key] = None
            return None
        tc = float(datos[-1]["dato"].replace(",", ""))
        st.session_state.bx_cache[cache_key] = tc
        return tc
    except Exception as exc:
        st.warning(f"Banxico ({moneda} · {fecha}): {exc}")
        st.session_state.bx_cache[cache_key] = None
        return None


# ─────────────────────────────────────────────────────────────
# RECALCULAR COLUMNAS DERIVADAS
# ─────────────────────────────────────────────────────────────
def recalc_derived(df: pd.DataFrame) -> pd.DataFrame:
    """Recalcula sub, iva y total de forma idempotente tras edición."""
    df = _normalize_editor_df(df)
    for idx, row in df.iterrows():
        qty = _safe_f(row.get("Cantidad"))
        qty = qty if qty and qty > 0 else 1

        unit = _safe_f(row.get("Precio Unitario"))
        sub  = _safe_f(row.get("Subtotal (Sin IVA)"))
        iva  = _safe_f(row.get("IVA 16%"))
        tot  = _safe_f(row.get("Total con IVA"))
        anx  = _safe_f(row.get("Monto en Anexo Escrito"))

        iva_flag = (
            str(row.get("(+ IVA)", "N/M") or "N/M").strip().lower()
            .replace("\u00ed", "i").replace("\u00c3\u00ad", "i")
        )
        has_iva = iva_flag in {"si", "yes", "true", "1"}
        no_iva  = iva_flag in {"no", "n/m", "nm", "n.a.", "na", ""}

        if sub is None and tot is not None and iva is not None:
            sub = round(tot - iva, 2)
        if unit is None:
            if sub is not None:
                unit = round(sub / qty, 2)
            elif tot is not None:
                base = (tot / 1.16) if has_iva else tot
                unit = round(base / qty, 2)
        if unit is not None:
            sub = round(qty * unit, 2)
        if sub is not None:
            if has_iva:
                iva = round(sub * 0.16, 2)
                tot = round(sub + iva, 2)
            elif no_iva:
                iva = 0.0
                tot = round(sub, 2)
            else:
                iva = round(iva, 2) if iva is not None else None
                tot = round(sub + (iva or 0), 2)

        df.at[idx, "Cantidad"]          = int(qty) if float(qty).is_integer() else qty
        df.at[idx, "Precio Unitario"]   = unit
        df.at[idx, "Subtotal (Sin IVA)"] = sub
        df.at[idx, "IVA 16%"]           = iva
        df.at[idx, "Total con IVA"]     = tot
        df.at[idx, "Diferencia final"]  = (
            round((tot or 0) - (anx or 0), 2)
            if tot is not None or anx is not None else None
        )
    return df


# ─────────────────────────────────────────────────────────────
# MOTOR DE EXTRACCIÓN
# ─────────────────────────────────────────────────────────────
def _parse_space_table(text: str) -> list[dict]:
    items = []
    for line in text.splitlines():
        m = _ITEM_RE.match(line.strip())
        if m:
            qty, desc, pu, total = m.groups()
            items.append({
                "qty": int(qty), "desc": desc.strip(),
                "pu": float(pu.replace(",", "")),
                "total": float(total.replace(",", "")),
            })
    return items


def extract(
    pdf_bytes: bytes,
    label: str,
    p0: int,
    p1: int,
    det_iva: bool,
    calc_sub: bool,
    tipo: str = "Cotización Proveedor",
    moneda: str = "MXN",
    bx_token: str = "",
) -> dict:
    """Motor de extracción de montos desde PDF (5 estrategias + cota proporcional)."""
    native_text = ""
    table_rows: list[list[str]] = []
    ocr_used = False

    # Estrategia 1 & 2: texto nativo + tablas
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        n_pages = len(pdf.pages)
        pr = range(max(0, p0 - 1), min(p1, n_pages))
        for i in pr:
            pg = pdf.pages[i]
            txt = pg.extract_text() or ""
            native_text += "\n" + txt
            for tbl in pg.extract_tables() or []:
                if tbl:
                    table_rows.extend(
                        [str(c or "").strip() for c in row] for row in tbl if row
                    )
            for item in _parse_space_table(txt):
                table_rows.append([
                    str(item["qty"]), item["desc"],
                    f"${item['pu']:,.2f}", f"${item['total']:,.2f}",
                ])

    pages_count = max(len(pr), 1)
    if len(native_text.strip()) < NATIVE_MIN_CHARS_PER_PAGE * pages_count:
        ocr_used = True
        for i in pr:
            o_txt = _ocr_page(pdf_bytes, i)
            if o_txt:
                native_text += "\n" + o_txt
                for item in _parse_space_table(o_txt):
                    table_rows.append([
                        str(item["qty"]), item["desc"],
                        f"${item['pu']:,.2f}", f"${item['total']:,.2f}",
                    ])

    text = native_text
    tlow = text.lower()

    iva_f = "Sí" if det_iva and re.search(r"\biva\b|16%|vat", tlow) else "N/M"
    tot = iva = sub = pu = fecha = None
    qty = 1
    obs_parts: list[str] = []
    if ocr_used:
        obs_parts.append("OCR")

    # ── Estrategia 1: tablas estructuradas ───────────────────
    for row in table_rows:
        j = "   ".join(row).lower()
        s = "   ".join(row)
        if re.search(r"\btotal\b", j) and not re.search(r"sub|parcial|acum", j):
            v = _money(s)
            if v and (tot is None or v > tot): tot = v
        if re.search(r"\biva\b|16%|vat", j):
            v = _money(s)
            if v and (iva is None or v > iva): iva = v
        if re.search(r"subtotal|sin\s*iva|importe\s*neto", j):
            v = _money(s)
            if v and (sub is None or v > sub): sub = v

    # ── Estrategia 2 (R1): triangulación qty × pu ≈ total ────
    valid_line_totals: list[float] = []
    seen_lines: set[tuple] = set()
    for row in table_rows:
        row_str = " ".join(str(c) for c in row)
        nums: list[float] = []
        for token in re.findall(r"[\d,]+(?:\.\d+)?", row_str):
            n = _safe_f(token.replace(",", ""))
            if n is not None and n > 0:
                nums.append(n)
        if len(nums) < 2:
            continue
        found = False
        for i_t, t_cand in enumerate(nums):
            if t_cand < 1.0 or found: continue
            for i_p, p_cand in enumerate(nums):
                if i_p == i_t or p_cand <= 0 or found: continue
                for i_q, q_cand in enumerate(nums):
                    if i_q in (i_t, i_p): continue
                    if not (1 <= q_cand <= 9999 and q_cand == int(q_cand)): continue
                    tol = max(0.5, t_cand * 0.01)
                    if abs(q_cand * p_cand - t_cand) <= tol:
                        key = (int(q_cand), round(p_cand, 2), round(t_cand, 2))
                        if key not in seen_lines:
                            seen_lines.add(key)
                            valid_line_totals.append(t_cand)
                        found = True
                        break
                if found: break

    if tot is None and valid_line_totals:
        tot = round(sum(valid_line_totals), 2)
        obs_parts.append("Total por líneas")

    # ── Estrategia 3: fecha y total por texto libre ───────────
    for ln in text.splitlines():
        if fecha is None:
            d = _date(ln)
            if d: fecha = d

    if tot is None:
        for ln in text.splitlines():
            if re.search(r"\btotal\b", ln, re.I) and not re.search(r"sub|parcial|acum", ln, re.I):
                v = _money(ln)
                if v and (tot is None or v > tot): tot = v

    # ── Estrategia 3.5: "es de $X" ────────────────────────────
    if tot is None:
        for ln in text.splitlines():
            m_ed = _ES_DE_RE.search(ln)
            if m_ed:
                v = _safe_f(m_ed.group(1).replace(",", ""))
                if v and 10.0 <= v <= MAX_PLAUSIBLE_MXN:
                    tot = v
                    obs_parts.append("Precio directo")
                    break

    # ── V1: Cota de Cordura Proporcional ─────────────────────
    # Descarta el total si es desproporcionado con respecto al subtotal
    # o a la suma de las líneas individuales encontradas.
    if tot is not None:
        if sub is not None and tot > sub * 1.5:
            obs_parts.append("⚠ Total descartado (desproporcionado vs subtotal)")
            tot = None
        elif valid_line_totals:
            sum_lines = sum(valid_line_totals)
            if sum_lines > 0 and tot > sum_lines * 1.5:
                obs_parts.append("⚠ Total descartado (desproporcionado vs líneas)")
                tot = None

    # ── Estrategia 4 (R2): fallback — máximo en contexto monetario ─
    if tot is None:
        all_vals: list[float] = []
        for ln in text.splitlines():
            if _BUDGET_EXCL.search(ln): continue  # V2: blacklist
            if not _MONETARY_CTX.search(ln): continue
            for m in _MONEY_RE.finditer(ln):
                raw = m.group(1) or m.group(2) or m.group(3)
                if raw:
                    try:
                        v = float(raw.replace(",", ""))
                        if 10.0 <= v <= MAX_PLAUSIBLE_MXN:
                            all_vals.append(v)
                    except ValueError:
                        pass
        if all_vals:
            # V3: también aplicar cota proporcional en el fallback
            if sub is not None:
                all_vals = [v for v in all_vals if v <= sub * 1.5]
            elif valid_line_totals:
                sum_lines = sum(valid_line_totals)
                all_vals = [v for v in all_vals if v <= sum_lines * 1.5]
            if all_vals:
                tot = max(all_vals)
                obs_parts.append("Total inferido (máx. validado)")

    # Límite global de cordura
    if tot is not None and tot > MAX_PLAUSIBLE_MXN:
        obs_parts.append(f"⚠ Valor sospechoso ({tot:,.2f}) — verificar")
        tot = None

    # Subtotal e IVA por texto libre
    if sub is None:
        for ln in text.splitlines():
            if re.search(r"subtotal|importe|sin\s*iva", ln, re.I) and \
               not re.search(r"\btotal\b", ln.replace("subtotal", ""), re.I):
                v = _money(ln)
                if v and (tot is None or v <= tot):
                    sub = v; break

    if iva is None and iva_f == "Sí":
        for ln in text.splitlines():
            if re.search(r"\biva\b|16%|vat|impuesto", ln, re.I):
                v = _money(ln)
                if v and (tot is None or v < tot):
                    iva = v; break

    # Cantidad y P.U. de tablas
    for row in table_rows:
        row_str = " ".join(str(c) for c in row)
        nums_: list[float] = []
        for t in re.findall(r"[\d,]+(?:\.\d+)?", row_str):
            n_val = _safe_f(t)
            if n_val is not None: nums_.append(n_val)
        if len(nums_) >= 2 and 1 <= nums_[0] <= 9999:
            qty = int(nums_[0])
            pu = nums_[-2] if len(nums_) > 2 else nums_[-1]

    # Triangulación final
    if tot and not sub and not iva and iva_f == "Sí":
        sub = round(tot / 1.16, 2)
        iva = round(tot - sub, 2)
        obs_parts.append("IVA desglosado")
    elif tot and iva and not sub:
        sub = round(tot - iva, 2)
    elif sub and iva and not tot:
        tot = round(sub + iva, 2)
    elif calc_sub and sub and not iva and not tot and iva_f == "Sí":
        iva = round(sub * 0.16, 2)
        tot = round(sub + iva, 2)

    if pu is None and sub is not None and qty > 0:
        pu = round(sub / qty, 2)

    # Conversión Banxico
    tc: Optional[float] = None
    if moneda != "MXN" and bx_token:
        fecha_tc = fecha or datetime.date.today()
        tc = _banxico_tc(moneda, fecha_tc, bx_token)
        if tc and tot is not None:
            tot_orig = tot
            tot = round(tot * tc, 2)
            sub = round(sub * tc, 2) if sub else None
            iva = round(iva * tc, 2) if iva else None
            pu  = round(pu  * tc, 2) if pu  else None
            obs_parts.append(f"1 {moneda} = ${tc:.4f} MXN")
            obs_parts.append(f"Total orig.: {moneda} {tot_orig:,.2f}")

    return {
        "Tipo":                 tipo,
        "Fecha":                fecha.isoformat() if fecha else datetime.date.today().isoformat(),
        "Rubro":                label,
        "QT":                   "Sí",
        "T. Cambio":            f"{moneda} ({tc:.4f})" if tc else moneda,
        "(+ IVA)":              iva_f,
        "Cantidad":             qty,
        "Precio Unitario":      pu,
        "Subtotal (Sin IVA)":   sub,
        "IVA 16%":              iva,
        "Total con IVA":        tot,
        "Diferencia final":     None,
        "Monto en Anexo Escrito": None,
        "Observaciones":        " | ".join(obs_parts) if obs_parts else "",
    }


# ─────────────────────────────────────────────────────────────
# EXPORTACIÓN EXCEL
# ─────────────────────────────────────────────────────────────
def _side(style: str = "thin") -> Side:
    return Side(style=style, color="000000")

def _border(style: str = "thin") -> Border:
    s = _side(style)
    return Border(left=s, right=s, top=s, bottom=s)

def _fill(rgb: str) -> PatternFill:
    return PatternFill("solid", start_color=rgb, end_color=rgb)


def to_excel(df: pd.DataFrame, nombre: str = "", blank: bool = False) -> bytes:
    """
    Genera Excel con formato profesional.
    blank=True → plantilla vacía con fórmulas vivas.
    """
    buf = io.BytesIO()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = nombre[:31] if nombre else "Conciliación"

    F_WHITE = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    F_HDR   = Font(name="Calibri", size=11, bold=True, color="000000")
    F_DATA  = Font(name="Calibri", size=11, color="000000")
    F_TOT   = Font(name="Calibri", size=11, bold=True, color="000000")
    F_BOLD  = Font(name="Calibri", size=11, bold=True, color="000000")

    FILL_ROW1 = _fill("6E152E")
    FILL_HDR_A = _fill("D4C19C")
    FILL_HDR_B = _fill("EBE2D1")
    BD     = _border("thin")
    BD_MED = _border("medium")
    AL_C = Alignment(horizontal="center", vertical="center", wrap_text=True)
    AL_L = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    AL_R = Alignment(horizontal="right",  vertical="center")

    n_rows = len(df)

    for col_idx, width in enumerate(_WIDTHS_COL, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Fila 1: nombre del proyecto
    ws.row_dimensions[1].height = 27.75
    c = ws.cell(row=1, column=1, value=nombre or "")
    c.font = F_WHITE; c.fill = FILL_ROW1; c.border = BD; c.alignment = AL_L
    for ci in range(2, len(_COLS) + 1):
        cx = ws.cell(row=1, column=ci)
        cx.fill = FILL_ROW1; cx.border = BD

    # Fila 2: encabezados
    ws.row_dimensions[2].height = 21.75
    for ci, hdr in enumerate(_COLS, 1):
        c = ws.cell(row=2, column=ci, value=hdr)
        c.font = F_HDR
        c.fill = FILL_HDR_A if ci <= 12 else FILL_HDR_B
        c.border = BD; c.alignment = AL_C

    DS = 3  # Data Start

    def _money_cell(ws_, r, ci, val=None):
        c_ = ws_.cell(row=r, column=ci, value=val)
        c_.number_format = _FMT_MONEY; c_.font = F_DATA
        c_.border = BD; c_.alignment = AL_R
        return c_

    for i, (_, row) in enumerate(df.iterrows()):
        r = DS + i
        ws.row_dimensions[r].height = 18

        # Col 1: Tipo
        c = ws.cell(row=r, column=1, value=str(row.get("Tipo", "")))
        c.font = F_DATA; c.border = BD; c.alignment = AL_C

        # Col 2: Fecha
        fv = row.get("Fecha")
        if isinstance(fv, str):
            try: fv = datetime.date.fromisoformat(fv)
            except Exception: fv = None
        if isinstance(fv, (datetime.date, datetime.datetime)):
            dt = datetime.datetime.combine(fv, datetime.time()) if isinstance(fv, datetime.date) else fv
            c = ws.cell(row=r, column=2, value=dt); c.number_format = _FMT_DATE
        else:
            c = ws.cell(row=r, column=2, value=str(fv or ""))
        c.font = F_DATA; c.border = BD; c.alignment = AL_C

        # Col 3: Rubro
        c = ws.cell(row=r, column=3, value=str(row.get("Rubro", "") or ""))
        c.font = F_DATA; c.border = BD; c.alignment = AL_L

        # Col 4: QT
        c = ws.cell(row=r, column=4, value=str(row.get("QT", "Sí")))
        c.font = F_DATA; c.border = BD; c.alignment = AL_C

        # Col 5: T. Cambio
        c = ws.cell(row=r, column=5, value=str(row.get("T. Cambio", "MXN") or "MXN"))
        c.font = F_DATA; c.border = BD; c.alignment = AL_C

        # Col 6: (+IVA)
        c = ws.cell(row=r, column=6, value=str(row.get("(+ IVA)", "") or ""))
        c.font = F_DATA; c.border = BD; c.alignment = AL_C

        # Col 7: Cantidad
        q = _safe_f(row.get("Cantidad"))
        c = ws.cell(row=r, column=7, value=int(q) if q is not None else 1)
        c.number_format = "0.00"; c.font = F_DATA; c.border = BD; c.alignment = AL_C

        # Col 8: Precio Unitario (input)
        pu_val = None if blank else _safe_f(row.get("Precio Unitario"))
        _money_cell(ws, r, 8, pu_val)

        has_iva = (
            str(row.get("(+ IVA)", "N/M")).strip().lower()
            .replace("\u00ed", "i").replace("\u00c3\u00ad", "i") == "si"
        )

        # Col 9: Subtotal = G × H  (col 7 × col 8)
        sub_val = f'=IF(ISNUMBER(H{r}),G{r}*H{r},"")' if blank else _safe_f(row.get("Subtotal (Sin IVA)"))
        c = ws.cell(row=r, column=9, value=sub_val)
        c.number_format = _FMT_MONEY; c.font = F_DATA; c.border = BD; c.alignment = AL_R

        # Col 10: IVA 16%
        if blank and has_iva:
            iva_val = f'=IF(ISNUMBER(I{r}),I{r}*0.16,"")'
        elif blank:
            iva_val = None
        else:
            iva_val = _safe_f(row.get("IVA 16%"))
        c = ws.cell(row=r, column=10, value=iva_val)
        c.number_format = _FMT_MONEY; c.font = F_DATA; c.border = BD; c.alignment = AL_R

        # Col 11: Total con IVA
        if blank and has_iva:
            tot_val = f'=IF(ISNUMBER(I{r}),IF(ISNUMBER(J{r}),I{r}+J{r},I{r}),"")'
        elif blank:
            tot_val = f'=IF(ISNUMBER(I{r}),I{r},"")'
        else:
            tot_val = _safe_f(row.get("Total con IVA"))
        c = ws.cell(row=r, column=11, value=tot_val)
        c.number_format = _FMT_MONEY; c.font = F_DATA; c.border = BD; c.alignment = AL_R

        # Col 12: Diferencia = K(11) - M(13)
        dif_val = f'=IF(AND(ISNUMBER(K{r}),ISNUMBER(M{r})),K{r}-M{r},"")' if blank else _safe_f(row.get("Diferencia final"))
        c = ws.cell(row=r, column=12, value=dif_val)
        c.number_format = _FMT_MONEY; c.font = F_DATA; c.border = BD; c.alignment = AL_R

        # Col 13: Monto en Anexo Escrito (input)
        anx_val = None if blank else _safe_f(row.get("Monto en Anexo Escrito"))
        c = _money_cell(ws, r, 13, anx_val); c.font = F_BOLD

        # Col 14: Observaciones
        c = ws.cell(row=r, column=14, value="" if blank else str(row.get("Observaciones", "") or ""))
        c.font = F_DATA; c.border = BD; c.alignment = AL_L

    # Fila de totales
    if n_rows > 0:
        tot_row = DS + n_rows
        s_xl, e_xl = DS, DS + n_rows - 1
        ws.row_dimensions[tot_row].height = 18
        c = ws.cell(row=tot_row, column=1, value="TOTALES")
        c.font = F_TOT; c.border = BD_MED
        # Sumatorias: Total(11), Diferencia(12), Anexo(13)
        for ci in (11, 12, 13):
            col_l = get_column_letter(ci)
            c = ws.cell(row=tot_row, column=ci, value=f"=SUM({col_l}{s_xl}:{col_l}{e_xl})")
            c.number_format = _FMT_MONEY; c.font = F_TOT
            c.fill = _fill("EBE2D1"); c.border = BD_MED; c.alignment = AL_R

    ws.freeze_panes = "A3"
    wb.save(buf)
    buf.seek(0)
    return buf.read()


# ─────────────────────────────────────────────────────────────
# CALLBACKS DE NAVEGACIÓN (N1)
# ─────────────────────────────────────────────────────────────
def _go_to(target: int) -> None:
    """Mueve el visor a la página indicada y sincroniza el widget numérico."""
    tp = max(st.session_state.total_pages - 1, 0)
    target = max(0, min(tp, target))
    st.session_state.current_page = target
    st.session_state.nav_page_input = target + 1

def _on_page_input() -> None:
    """Callback del number_input de página; actualiza current_page."""
    target = st.session_state.nav_page_input - 1
    tp = max(st.session_state.total_pages - 1, 0)
    st.session_state.current_page = max(0, min(tp, target))


# ─────────────────────────────────────────────────────────────
# SIDEBAR
# ─────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## 📂 Documentos PDF")
    uploaded = st.file_uploader(
        "Sube uno o varios PDFs de cotizaciones",
        type=["pdf"],
        accept_multiple_files=True,
    )

    if uploaded:
        new_files: list[dict] = []
        for f in uploaded:
            raw = f.getvalue()
            h = _md5(raw)
            with fitz.open(stream=raw, filetype="pdf") as doc:
                n = len(doc)
            new_files.append({"name": f.name, "bytes": raw, "hash": h, "pages": n})

        # V5: usamos pdf_upload_hash para comparar el conjunto subido
        new_upload_hash = _md5("".join(fi["hash"] for fi in new_files).encode())
        if new_upload_hash != st.session_state.pdf_upload_hash:
            combined_bytes, combined_hash, total_pages, page_map = _build_combined_pdf(new_files)
            st.session_state.update(
                pdf_files=new_files,
                pdf_combined=combined_bytes,
                pdf_combined_hash=combined_hash,
                pdf_upload_hash=new_upload_hash,
                total_pages=total_pages,
                page_map=page_map,
                current_page=0,
                nav_page_input=1,
                extracted=False,
                df=None,
                editor_df=None,
                df_hash="",
            )
            st.session_state.editor_version += 1

        n_files = len(st.session_state.pdf_files)
        tp_total = st.session_state.total_pages
        if n_files == 1:
            st.success(f"✅ 1 archivo · {tp_total} págs.")
        else:
            st.success(f"✅ {n_files} archivos · {tp_total} págs. totales")
            with st.expander("📄 Detalle de archivos"):
                for idx, fi in enumerate(st.session_state.pdf_files):
                    st.caption(f"**{idx + 1}.** {fi['name']}  —  {fi['pages']} pág(s).")

    st.markdown("---")
    st.markdown("### 🏷 Proyecto")
    st.session_state.proyecto_nombre = st.text_input(
        "Nombre del proyecto", value=st.session_state.proyecto_nombre
    )

    st.markdown("---")
    st.markdown("### 💱 Banxico – Tipo de Cambio")
    st.markdown(
        "Obtén tu token gratuito en [**SIE Banxico API** →](https://www.banxico.org.mx/SieAPIRest/) "
        "*(solo si hay cotizaciones en USD/EUR/CAD)*"
    )
    _token_default = st.secrets.get("BANXICO_TOKEN", "") if hasattr(st, "secrets") else ""
    bx_token = st.text_input(
        "Token Bmx-Token", value=_token_default, type="password",
        placeholder="Pega aquí tu token",
    )
    if bx_token:
        st.caption("🔑 Token activo")

    st.markdown("---")
    st.markdown("### ⚙️ Secciones (cotizaciones)")
    n = int(st.number_input(
        "Número de secciones", min_value=1, max_value=50,
        value=st.session_state.num_sec, step=1,
    ))
    if n != st.session_state.num_sec:
        st.session_state.num_sec = n
        st.session_state.extracted = False
        st.session_state.df = None
        st.session_state.editor_df = None
        st.session_state.df_hash = ""
        st.session_state.editor_version += 1

    cfgs = list(st.session_state.sec_cfg)
    pdf_names = [fi["name"] for fi in st.session_state.pdf_files] if st.session_state.pdf_files else []

    while len(cfgs) < n:
        i = len(cfgs) + 1
        cfgs.append({
            "label":    f"Sección {i}",
            "tipo":     "Cotización Proveedor",
            "p0": 1, "p1": 1,
            "det_iva":  True,
            "calc_sub": True,
            "moneda":   "MXN",
            "pdf_idx":  0,
        })
    cfgs = cfgs[:n]

    for i, c in enumerate(cfgs):
        with st.expander(f"📄 {c['label']}", expanded=(n <= 6)):
            c["label"] = st.text_input("Rubro / Concepto", value=c["label"], key=f"lb{i}")

            # N3: selector de tipo
            c["tipo"] = st.selectbox(
                "Tipo",
                ["Cotización Proveedor", "Presupuesto Global"],
                index=0 if c.get("tipo", "Cotización Proveedor") == "Cotización Proveedor" else 1,
                key=f"tipo{i}",
            )

            if len(pdf_names) > 1:
                safe_idx = min(c.get("pdf_idx", 0), len(pdf_names) - 1)
                c["pdf_idx"] = st.selectbox(
                    "PDF fuente",
                    range(len(pdf_names)),
                    index=safe_idx,
                    format_func=lambda x: pdf_names[x],
                    key=f"pdf{i}",
                )
            else:
                c["pdf_idx"] = 0

            tp_local = max(st.session_state.pdf_files[c["pdf_idx"]]["pages"], 1) if st.session_state.pdf_files else 1

            ca, cb = st.columns(2)
            c["p0"] = ca.number_input("Pág. Inicio", 1, tp_local, min(c["p0"], tp_local), key=f"p0{i}")
            c["p1"] = cb.number_input("Pág. Fin", c["p0"], tp_local, max(min(c["p1"], tp_local), c["p0"]), key=f"p1{i}")

            c["moneda"] = st.selectbox(
                "Moneda",
                ["MXN", "USD", "EUR", "CAD"],
                index=["MXN", "USD", "EUR", "CAD"].index(c.get("moneda", "MXN")),
                key=f"mon{i}",
            )
            c["det_iva"]  = st.checkbox("Detectar IVA", value=c["det_iva"], key=f"iv{i}")
            c["calc_sub"] = st.checkbox("Calcular subtotal si falta", value=c.get("calc_sub", True), key=f"cs{i}")

    st.session_state.sec_cfg = cfgs
    st.markdown("---")
    run = st.button(
        "🔍 Extraer Montos",
        disabled=(st.session_state.pdf_combined is None),
        use_container_width=True,
        type="primary",
    )


# ─────────────────────────────────────────────────────────────
# CABECERA PRINCIPAL
# ─────────────────────────────────────────────────────────────
st.markdown(
    '<div class="hdr"><h1>📋 Conciliador de Cotizaciones</h1>'
    '<p>Extracción automática PDF · Carga múltiple · Cota de Cordura · v5.0</p></div>',
    unsafe_allow_html=True,
)


# ─────────────────────────────────────────────────────────────
# CONTROLADOR DE EXTRACCIÓN
# ─────────────────────────────────────────────────────────────
if run and st.session_state.pdf_combined:
    rows: list[dict] = []
    n_secs = st.session_state.num_sec
    bar = st.progress(0, text="Iniciando extracción…")

    for i, c in enumerate(st.session_state.sec_cfg):
        bar.progress((i + 0.3) / n_secs, text=f"🔍 {c['label']}…")
        pdf_idx = c.get("pdf_idx", 0)
        section_pdf_bytes = (
            st.session_state.pdf_files[pdf_idx]["bytes"]
            if pdf_idx < len(st.session_state.pdf_files)
            else st.session_state.pdf_combined
        )
        try:
            row = extract(
                section_pdf_bytes,
                c["label"],
                c["p0"],
                c["p1"],
                c["det_iva"],
                c.get("calc_sub", True),
                tipo=c.get("tipo", "Cotización Proveedor"),
                moneda=c.get("moneda", "MXN"),
                bx_token=bx_token,
            )
            rows.append(row)
        except Exception as exc:
            st.warning(f"⚠️ Sección {i + 1} «{c['label']}»: {str(exc)[:120]}")
            rows.append({
                **{k: None for k in _COLS},
                "Tipo":  c.get("tipo", "Cotización Proveedor"),
                "Rubro": c["label"],
                "QT":    "Sí",
                "T. Cambio": c.get("moneda", "MXN"),
                "Cantidad": 1,
                "Fecha": datetime.date.today().isoformat(),
                "Observaciones": f"Error: {str(exc)[:100]}",
            })
        bar.progress((i + 1) / n_secs)

    bar.empty()
    df_new = pd.DataFrame(rows).reindex(columns=_COLS)
    df_new = recalc_derived(df_new)
    st.session_state.df = df_new
    st.session_state.df_hash = _df_hash(df_new)
    st.session_state.extracted = True
    st.session_state._rerun_guard = False
    st.success(f"✅ {len(rows)} sección(es) procesada(s).")


if st.session_state.pdf_combined is None:
    st.info("Sube uno o varios PDFs en la barra lateral para comenzar.")
    st.stop()


# ─────────────────────────────────────────────────────────────
# LAYOUT: VISOR + EDITOR
# ─────────────────────────────────────────────────────────────
col_L, col_R = st.columns(2, gap="medium")


# ── VISOR ────────────────────────────────────────────────────
with col_L:
    st.markdown('<p class="ptitle">🔍 Visor de Documento</p>', unsafe_allow_html=True)
    tp = st.session_state.total_pages

    nav_p, nav_c, nav_n = st.columns([1, 4, 1])
    nav_p.button(
        "◀", key="btn_prev", use_container_width=True,
        on_click=_go_to, args=(st.session_state.current_page - 1,),
    )
    nav_n.button(
        "▶", key="btn_next", use_container_width=True,
        on_click=_go_to, args=(st.session_state.current_page + 1,),
    )
    # N1: number_input vinculado a nav_page_input; on_change sincroniza current_page
    nav_c.number_input(
        "Página", min_value=1, max_value=tp, step=1,
        label_visibility="collapsed",
        key="nav_page_input",
        on_change=_on_page_input,
    )

    cp = st.session_state.current_page
    st.caption(f"Página {cp + 1} de {tp}")

    pm = st.session_state.page_map
    if pm and cp < len(pm):
        pg_info = pm[cp]
        st.markdown(
            f'<span class="tag-file">📁 {pg_info["name"]}</span> '
            f'<span style="font-size:.8rem;color:var(--app-muted)">'
            f'pág. local {pg_info["local_page"] + 1}</span>',
            unsafe_allow_html=True,
        )

    # Indicador de sección activa
    for c_cfg in st.session_state.sec_cfg:
        pdf_idx = c_cfg.get("pdf_idx", 0)
        global_offset = sum(
            fi["pages"] for fi in st.session_state.pdf_files[:pdf_idx]
        ) if st.session_state.pdf_files else 0
        g_p0 = global_offset + c_cfg["p0"]
        g_p1 = global_offset + c_cfg["p1"]
        if g_p0 <= cp + 1 <= g_p1:
            tipo_badge = (
                "🌐 Presupuesto Global"
                if c_cfg.get("tipo") == "Presupuesto Global"
                else f"📑 {c_cfg['label']}"
            )
            st.markdown(
                f'<span style="background:var(--accent);color:#fff;'
                f'padding:3px 10px;border-radius:4px;font-size:.8rem">'
                f'{tipo_badge} · <span class="tag-moneda">'
                f'{c_cfg.get("moneda","MXN")}</span></span>',
                unsafe_allow_html=True,
            )
            break

    with st.spinner("Cargando…"):
        st.image(_render(st.session_state.pdf_combined_hash, cp), use_container_width=True)

    # Descarga(s) de PDF
    if len(st.session_state.pdf_files) == 1:
        st.download_button(
            "📥 Descargar PDF",
            data=st.session_state.pdf_files[0]["bytes"],
            file_name=st.session_state.pdf_files[0]["name"],
            mime="application/pdf",
            use_container_width=True,
        )
    elif len(st.session_state.pdf_files) > 1:
        st.download_button(
            "📥 Descargar PDF consolidado",
            data=st.session_state.pdf_combined,
            file_name="cotizaciones_consolidado.pdf",
            mime="application/pdf",
            use_container_width=True,
        )
        with st.expander("📄 Descargar archivos individuales"):
            for idx, fi in enumerate(st.session_state.pdf_files):
                st.download_button(
                    f"📥 {fi['name']}", data=fi["bytes"],
                    file_name=fi["name"], mime="application/pdf",
                    use_container_width=True, key=f"dl_pdf_{idx}",
                )


# ── EDITOR DE DATOS ──────────────────────────────────────────
with col_R:
    st.markdown('<p class="ptitle">✏️ Editor de Datos</p>', unsafe_allow_html=True)

    if not st.session_state.extracted or st.session_state.df is None:
        st.info("Configura las secciones y presiona **🔍 Extraer Montos**.")
    else:
        df_cur = st.session_state.df
        if df_cur is None or df_cur.empty:
            st.warning("Sin datos. Ejecuta nuevamente la extracción.")
            st.stop()

        kpi_slot = st.container()

        edited_df = st.data_editor(
            df_cur,
            key="data_editor_main",
            use_container_width=True,
            hide_index=True,
            num_rows="dynamic",
            column_config={
                "Tipo": st.column_config.SelectboxColumn(
                    "Tipo",
                    options=["Cotización Proveedor", "Presupuesto Global"],
                    width="medium",
                ),
                "Fecha":    st.column_config.TextColumn("Fecha", width="small"),
                "Rubro":    st.column_config.TextColumn("Rubro", width="large"),
                "QT":       st.column_config.SelectboxColumn("QT", options=["Sí", "No"], width="small"),
                "T. Cambio": st.column_config.TextColumn("Moneda/TC", width="small"),
                "(+ IVA)":  st.column_config.SelectboxColumn("IVA", options=["Sí", "No", "N/M"], width="small"),
                "Cantidad": st.column_config.NumberColumn("Cant.", format="%d", width="small"),
                "Precio Unitario":    st.column_config.NumberColumn("P. Unit.",  format="$%.2f", width="medium"),
                "Subtotal (Sin IVA)": st.column_config.NumberColumn("Subtotal",  format="$%.2f", width="medium"),
                "IVA 16%":            st.column_config.NumberColumn("IVA 16%",   format="$%.2f", width="medium"),
                "Total con IVA":      st.column_config.NumberColumn("Total c/IVA", format="$%.2f", width="medium"),
                "Diferencia final":   st.column_config.NumberColumn("Diferencia", format="$%.2f", width="medium", disabled=True),
                "Monto en Anexo Escrito": st.column_config.NumberColumn("Anexo $", format="$%.2f", width="medium"),
                "Observaciones": st.column_config.TextColumn("Observaciones", width="large"),
            },
        )

        updated_df = recalc_derived(edited_df)
        new_hash = _df_hash(updated_df)
        if new_hash != st.session_state.df_hash:
            st.session_state.df = updated_df
            st.session_state.df_hash = new_hash
            if not st.session_state.get("_rerun_guard"):
                st.session_state._rerun_guard = True
                st.rerun()
            else:
                st.session_state._rerun_guard = False
        else:
            st.session_state._rerun_guard = False

        df_cur = st.session_state.df

        # ── KPIs dinámicos (N3: Presupuesto Global) ──────────
        df_prov   = df_cur[df_cur["Tipo"] == "Cotización Proveedor"]
        df_presup = df_cur[df_cur["Tipo"] == "Presupuesto Global"]

        ts_ = pd.to_numeric(df_prov["Total con IVA"], errors="coerce").sum()

        if not df_presup.empty:
            rs_ = pd.to_numeric(df_presup["Total con IVA"], errors="coerce").sum()
            kpi_lbl_2 = "Ppto. Global Extraído"
        else:
            rs_ = pd.to_numeric(df_cur["Monto en Anexo Escrito"], errors="coerce").sum()
            kpi_lbl_2 = "Monto Anexo Manual"

        dif = ts_ - rs_

        with kpi_slot:
            k1, k2, k3 = st.columns(3)
            for kol, val, lbl in [
                (k1, ts_, "Total Proveedores"),
                (k2, rs_, kpi_lbl_2),
                (k3, dif, "Diferencia"),
            ]:
                color = "var(--accent)" if lbl != "Diferencia" else ("#c0392b" if dif > 0.01 else "#28a745")
                kol.markdown(
                    f'<div class="kpi"><div class="v" style="color:{color}">'
                    f"${val:,.2f}</div><div class=\"l\">{lbl}</div></div>",
                    unsafe_allow_html=True,
                )

        warn_rows = df_cur[df_cur["Observaciones"].str.contains("⚠|OCR|inferido", na=False)]
        if not warn_rows.empty:
            with st.expander(f"⚠ {len(warn_rows)} aviso(s) de extracción"):
                for _, wr in warn_rows.iterrows():
                    st.caption(f"• **{wr['Rubro']}**: {wr['Observaciones']}")

        st.markdown("---")

        ts_str = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        pname  = (st.session_state.proyecto_nombre or "Cotizaciones").replace(" ", "_")
        nombre = st.session_state.proyecto_nombre

        try:
            xlsx_bytes = to_excel(df_cur, nombre=nombre)
            st.download_button(
                "⬇️ Descargar Excel",
                data=xlsx_bytes,
                file_name=f"Conciliacion_{pname}_{ts_str}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary",
            )
        except Exception as exc:
            st.error(f"Error generando Excel: {exc}")

        with st.expander("📝 Plantilla vacía para captura manual"):
            st.info(
                "Descarga la plantilla vacía: al ingresar el Precio Unitario, "
                "Excel calcula Subtotal, IVA y Total automáticamente."
            )
            n_sec = st.session_state.num_sec
            df_blank = pd.DataFrame({
                "Tipo":  [c.get("tipo", "Cotización Proveedor") for c in st.session_state.sec_cfg],
                "Fecha": [datetime.date.today().isoformat()] * n_sec,
                "Rubro": [c["label"] for c in st.session_state.sec_cfg],
                "QT":    ["Sí"] * n_sec,
                "T. Cambio": [c.get("moneda", "MXN") for c in st.session_state.sec_cfg],
                "(+ IVA)":   ["Sí"] * n_sec,
                "Cantidad":  [1] * n_sec,
                "Precio Unitario":    [None] * n_sec,
                "Subtotal (Sin IVA)": [None] * n_sec,
                "IVA 16%":            [None] * n_sec,
                "Total con IVA":      [None] * n_sec,
                "Diferencia final":   [None] * n_sec,
                "Monto en Anexo Escrito": [None] * n_sec,
                "Observaciones":      [""] * n_sec,
            })
            try:
                xlsx_blank = to_excel(df_blank, nombre=nombre, blank=True)
                st.download_button(
                    "📄 Descargar plantilla vacía",
                    data=xlsx_blank,
                    file_name=f"Plantilla_{pname}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as exc:
                st.error(f"Error generando plantilla: {exc}")
